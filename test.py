from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = "empty_book.xlsx"

ws = wb.active
ws.title = "range names"

equipes = [
    {"Fermat": 2, "Lautrec": 2, "Mirepoix": 6},
    {"Fermat": 2, "Lautrec": 3, "Mirepoix": 4},
]

for college in arrayColleges:
    ws.append([college.nom])
    for equipe in arrayEquipes:

        for key, value in equipes[i].items():
            ws.append([key, value])

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3["AA10"].value)

wb.save(filename=dest_filename)
