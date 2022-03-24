import numpy as np
from openpyxl import load_workbook
import math
from modelisation import College, Equipe


def init(arrayColleges, nbEquipes):
    arrayEquipes = [Equipe(f"Equipe_{i}", arrayColleges) for i in range(nbEquipes)]

    nbElevesTotal = sum([col.nbEleves for col in arrayColleges])
    maxParEquipe = math.ceil(nbElevesTotal / nbEquipes)

    return arrayEquipes, maxParEquipe


def readExcel():
    wb = load_workbook("./effectifs.xlsx", data_only=True)
    sheet = wb[wb.sheetnames[0]]
    noms = [cell.value for cell in sheet["A"][1:]]
    eleves = [cell.value for cell in sheet["B"][1:]]

    assert len(noms) == len(eleves)

    arrayColleges = [College(noms[i], eleves[i]) for i in range(len(noms))]
    arrayColleges.sort()
    return arrayColleges
